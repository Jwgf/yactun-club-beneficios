import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import urlopen

from openpyxl import Workbook


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
    CONFIG_PATH = BASE_DIR / "config.local.json"
else:
    BASE_DIR = Path(__file__).resolve().parent.parent
    CONFIG_PATH = BASE_DIR / "pc_informes" / "config.local.json"

OUT_DIR = BASE_DIR / "informes"


def cargar_config():
    with CONFIG_PATH.open("r", encoding="utf-8-sig") as f:
        cfg = json.load(f)

    api_url = str(cfg.get("api_url", "")).strip()
    clave = str(cfg.get("claveInforme", "")).strip()

    if not api_url:
        raise ValueError("Falta api_url en config.local.json")

    if not clave or clave == "PEGAR_AQUI_LA_CLAVE_PRIVADA":
        raise ValueError("Falta claveInforme real en config.local.json")

    return api_url, clave


def pedir_informe(api_url, clave, periodo):
    params = {
        "action": "informeMensual",
        "periodo": periodo,
        "clave": clave,
    }

    url = api_url + "?" + urlencode(params)

    with urlopen(url, timeout=45) as r:
        raw = r.read().decode("utf-8")

    data = json.loads(raw)

    if not data.get("ok"):
        raise RuntimeError(f"Error backend: {data.get('error')} - {data.get('mensaje')}")

    return data


def escribir_tabla(ws, filas, columnas):
    ws.append(columnas)

    for fila in filas:
        ws.append([fila.get(col, "") for col in columnas])

    for col_cells in ws.columns:
        max_len = 10
        letra = col_cells[0].column_letter

        for cell in col_cells:
            valor = "" if cell.value is None else str(cell.value)
            if len(valor) > max_len:
                max_len = min(len(valor), 45)

        ws.column_dimensions[letra].width = max_len + 2


def crear_excel(data, periodo):
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"

    resumen = data.get("resumen", {})

    ws.append(["Concepto", "Valor"])
    ws.append(["Periodo", data.get("periodo", periodo)])
    ws.append(["Generado", data.get("generadoEn", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))])
    ws.append(["Clientes nuevos", resumen.get("clientesNuevos", 0)])
    ws.append(["Clientes activos del mes", resumen.get("clientesActivosDelMes", 0)])
    ws.append(["Movimientos del mes", resumen.get("movimientosDelMes", 0)])
    ws.append(["Compras registradas", resumen.get("comprasRegistradas", 0)])
    ws.append(["Premios generados", resumen.get("premiosGenerados", 0)])
    ws.append(["Premios entregados", resumen.get("premiosEntregados", 0)])
    ws.append(["Premios pendientes totales", resumen.get("premiosPendientesTotales", 0)])

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    clientes_cols = [
        "clienteId", "codigo", "nombre", "celular", "email", "fechaAlta",
        "puntosActuales", "ciclosGanados", "premiosPendientes", "estado", "ultimaCompra"
    ]

    movimientos_cols = [
        "fecha", "clienteId", "codigo", "tipo", "puntosAntes", "puntosDespues", "vendedor", "observacion"
    ]

    premios_cols = [
        "fechaGanado", "clienteId", "codigo", "premio", "estado", "fechaEntregado", "vendedor"
    ]

    ws_clientes = wb.create_sheet("Clientes")
    escribir_tabla(ws_clientes, data.get("clientes", []), clientes_cols)

    ws_movimientos = wb.create_sheet("Movimientos")
    escribir_tabla(ws_movimientos, data.get("movimientos", []), movimientos_cols)

    ws_premios = wb.create_sheet("Premios")
    escribir_tabla(ws_premios, data.get("premios", []), premios_cols)

    nombre = f"Yactun_Informe_{periodo.replace('-', '_')}.xlsx"
    salida = OUT_DIR / nombre

    wb.save(salida)
    return salida


def main():
    parser = argparse.ArgumentParser(description="Generador de informes mensuales Yactun")
    parser.add_argument("--periodo", required=True, help="Periodo en formato YYYY-MM. Ejemplo: 2026-06")
    args = parser.parse_args()

    api_url, clave = cargar_config()
    data = pedir_informe(api_url, clave, args.periodo)
    salida = crear_excel(data, args.periodo)

    print("Informe creado correctamente:")
    print(salida)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ERROR:", e)
        sys.exit(1)
